import sys
import os
import re
import pymysql
import openpyxl

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))


def load_env():
    env_path = os.path.join(PROJECT_ROOT, '.env')
    if not os.path.exists(env_path):
        print('[ERROR] .env 文件不存在，请先在项目根目录创建 .env 文件')
        sys.exit(1)
    env = {}
    with open(env_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, _, value = line.partition('=')
                env[key.strip()] = value.strip()
    return env


def parse_db_uri(uri):
    m = re.match(r'mysql\+pymysql://([^:]+):([^@]+)@([^:]+):(\d+)/([^?]+)', uri)
    if not m:
        m = re.match(r'mysql\+pymysql://([^:]+):([^@]+)@([^/]+)/([^?]+)', uri)
        if not m:
            print(f'[ERROR] 无法解析 DATABASE_URI: {uri}')
            sys.exit(1)
        user, pwd, host, db = m.groups()
        port = 3306
    else:
        user, pwd, host, port, db = m.groups()
        port = int(port)
    return host, port, user, pwd, db


SHEET_TABLE_MAP = {
    '品名': 'product_name',
    '牌号': 'brand',
    '产地': 'origin',
    '仓库': 'warehouse',
    '部门': 'department',
    '单位名称': 'unit',
}

TABLES_SIMPLE = {'product_name', 'brand', 'origin', 'warehouse', 'department'}
TABLES_COMPLEX = {'unit'}


def import_simple_sheet(cursor, conn, sheet_name, table_name, ws):
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return 0, 0
    header = [str(c or '').strip() for c in rows[0]]
    name_col = None
    for i, h in enumerate(header):
        if h.lower() == 'name':
            name_col = i
            break
    if name_col is None:
        print(f'  [WARN] Sheet "{sheet_name}" 缺少 name 列，跳过')
        return 0, 0

    inserted = 0
    skipped = 0
    for row in rows[1:]:
        name = str(row[name_col] or '').strip()
        if not name:
            continue
        cursor.execute(f"SELECT id FROM {table_name} WHERE name = %s", (name,))
        if cursor.fetchone():
            skipped += 1
            continue
        try:
            cursor.execute(f"INSERT INTO {table_name} (name) VALUES (%s)", (name,))
            conn.commit()
            inserted += 1
        except Exception as e:
            conn.rollback()
            print(f'  [ERR] 插入 {table_name} "{name}" 失败: {e}')
    return inserted, skipped


def import_complex_sheet(cursor, conn, sheet_name, table_name, ws):
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return 0, 0
    header = [str(c or '').strip() for c in rows[0]]
    col_map = {}
    for i, h in enumerate(header):
        col_map[h.lower()] = i

    if 'name' not in col_map:
        print(f'  [WARN] Sheet "{sheet_name}" 缺少 name 列，跳过')
        return 0, 0

    inserted = 0
    skipped = 0
    for row in rows[1:]:
        name = str(row[col_map['name']] or '').strip()
        if not name:
            continue
        cursor.execute(f"SELECT id FROM {table_name} WHERE name = %s", (name,))
        if cursor.fetchone():
            skipped += 1
            continue
        contact = str(row[col_map['contact']] or '').strip() if 'contact' in col_map else ''
        phone = str(row[col_map['phone']] or '').strip() if 'phone' in col_map else ''
        remark = str(row[col_map['remark']] or '').strip() if 'remark' in col_map else ''
        try:
            cursor.execute(
                f"INSERT INTO {table_name} (name, contact, phone, remark) VALUES (%s, %s, %s, %s)",
                (name, contact, phone, remark)
            )
            conn.commit()
            inserted += 1
        except Exception as e:
            conn.rollback()
            print(f'  [ERR] 插入 {table_name} "{name}" 失败: {e}')
    return inserted, skipped


def main():
    if len(sys.argv) < 2:
        print('用法: python import_base_data.py <Excel文件路径>')
        print('示例: python import_base_data.py 基础数据.xlsx')
        sys.exit(1)

    excel_file = sys.argv[1]
    if not os.path.exists(excel_file):
        print(f'[ERROR] 文件不存在: {excel_file}')
        sys.exit(1)

    env = load_env()
    db_uri = env.get('DATABASE_URI')
    if not db_uri:
        print('[ERROR] .env 中未找到 DATABASE_URI')
        sys.exit(1)

    host, port, user, pwd, db_name = parse_db_uri(db_uri)
    print(f'数据库连接: {host}:{port}/{db_name}')

    try:
        conn = pymysql.connect(
            host=host, port=port, user=user, password=pwd,
            database=db_name, charset='utf8mb4'
        )
        cursor = conn.cursor()
    except Exception as e:
        print(f'[ERROR] 数据库连接失败: {e}')
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as e:
        print(f'[ERROR] 无法打开 Excel 文件: {e}')
        conn.close()
        sys.exit(1)

    print(f'\n开始导入 {os.path.basename(excel_file)}，共 {len(wb.sheetnames)} 个 Sheet\n')

    total_inserted = 0
    total_skipped = 0

    for sheet_name in wb.sheetnames:
        table_name = SHEET_TABLE_MAP.get(sheet_name)
        if not table_name:
            print(f'[SKIP] Sheet "{sheet_name}" 不在映射表中，跳过')
            continue

        ws = wb[sheet_name]
        print(f'[{sheet_name}] 正在导入...')

        if table_name in TABLES_SIMPLE:
            ins, skp = import_simple_sheet(cursor, conn, sheet_name, table_name, ws)
        elif table_name in TABLES_COMPLEX:
            ins, skp = import_complex_sheet(cursor, conn, sheet_name, table_name, ws)
        else:
            continue

        print(f'  新增 {ins} 条，跳过 {skp} 条重复')
        total_inserted += ins
        total_skipped += skp

    print(f'\n=== 导入完成 ===')
    print(f'总计新增: {total_inserted} 条')
    print(f'总计跳过: {total_skipped} 条重复')

    cursor.close()
    conn.close()


if __name__ == '__main__':
    main()
