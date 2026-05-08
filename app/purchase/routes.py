from flask import render_template, request, redirect, url_for, flash, current_app, send_file, session, jsonify
from flask_login import login_required, current_user
from app.purchase import purchase_bp
from app.purchase.services import create_purchase_order
from app.models import PurchaseOrder, PurchaseItem, Unit, Department, ProductName, Brand, Origin, Inventory, TransactionLog, SalesItem, ProcessDetail, Warehouse
from app import db
from datetime import date as date_type, timedelta
import openpyxl
from io import BytesIO
from itertools import groupby


@purchase_bp.route('/')
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    supplier_id = request.args.get('supplier_id', 0, type=int)
    dept_id = request.args.get('dept_id', 0, type=int)
    remark = request.args.get('remark', '').strip()
    search_submitted = request.args.get('search_submitted', '')

    default_date_start = (date_type.today() - timedelta(days=30)).isoformat()
    default_date_end = date_type.today().isoformat()

    if not search_submitted:
        units = Unit.query.all()
        departments = Department.query.all()
        return render_template('purchase/list.html', pagination=None, suppliers=units, departments=departments,
                               search_submitted=False, default_date_start=default_date_start, default_date_end=default_date_end)

    query = PurchaseOrder.query
    if date_start:
        query = query.filter(PurchaseOrder.order_date >= date_start)
    if date_end:
        query = query.filter(PurchaseOrder.order_date <= date_end)
    if supplier_id:
        query = query.filter(PurchaseOrder.supplier_id == supplier_id)
    if dept_id:
        query = query.filter(PurchaseOrder.dept_id == dept_id)
    if remark:
        query = query.filter(PurchaseOrder.remark.contains(remark))

    pagination = query.order_by(PurchaseOrder.order_date.desc(), PurchaseOrder.created_at.desc()).paginate(
        page=page, per_page=current_app.config['PER_PAGE'], error_out=False
    )
    units = Unit.query.all()
    departments = Department.query.all()
    total_qty_sum = sum(o.total_qty for o in pagination.items)
    total_weight_sum = sum(float(o.total_weight) for o in pagination.items)
    total_amount_sum = sum(float(o.total_amount) for o in pagination.items)
    return render_template('purchase/list.html', pagination=pagination, suppliers=units, departments=departments,
                           total_qty_sum=total_qty_sum, total_weight_sum=total_weight_sum, total_amount_sum=total_amount_sum,
                           default_date_start=default_date_start, default_date_end=default_date_end)


@purchase_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    if request.method == 'GET':
        units = Unit.query.all()
        departments = Department.query.all()
        product_names = ProductName.query.all()
        brands = Brand.query.all()
        origins = Origin.query.all()
        warehouses = Warehouse.query.all()
        return render_template('purchase/form.html', suppliers=units, departments=departments, product_names=product_names, brands=brands, origins=origins, warehouses=warehouses, today=date_type.today().isoformat())

    supplier_id = request.form.get('supplier_id', 0, type=int)
    dept_id = request.form.get('dept_id', 0, type=int)
    remark = request.form.get('remark', '')
    order_date_str = request.form.get('order_date', '')
    order_date_val = date_type.fromisoformat(order_date_str) if order_date_str else date_type.today()

    card_nos = request.form.getlist('card_no[]')
    product_names = request.form.getlist('product_name[]')
    brands = request.form.getlist('brand[]')
    origins = request.form.getlist('origin[]')
    specs = request.form.getlist('spec[]')
    qtys = request.form.getlist('qty[]')
    weights = request.form.getlist('weight[]')
    unit_prices = request.form.getlist('unit_price[]')
    remarks = request.form.getlist('remark[]')
    warehouses = request.form.getlist('warehouse[]')
    _wh_id = request.form.get('warehouse_id', type=int)

    items_data = []
    all_warehouses = {w.id: w.name for w in Warehouse.query.all()}
    wh_name = all_warehouses.get(_wh_id, '') if _wh_id else ''
    for i in range(len(card_nos)):
        items_data.append({
            'card_no': card_nos[i],
            'product_name': product_names[i],
            'brand': brands[i],
            'origin': origins[i],
            'spec': specs[i],
            'qty': int(qtys[i]),
            'weight': float(weights[i]),
            'unit_price': float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else 0,
            'warehouse': wh_name,
            'warehouse_id': _wh_id,
            'remark': remarks[i] if i < len(remarks) else ''
        })

    order_data = {
        'supplier_id': supplier_id,
        'dept_id': dept_id,
        'remark': remark,
        'order_date': order_date_val,
        'operator_id': current_user.id
    }
    create_purchase_order(order_data, items_data)
    flash('入库单创建成功', 'success')
    return redirect(url_for('purchase.index'))


@purchase_bp.route('/detail/<int:id>')
@login_required
def detail(id):
    order = PurchaseOrder.query.get_or_404(id)
    return render_template('purchase/form.html', order=order, readonly=True)


@purchase_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit(id):
    order = PurchaseOrder.query.get_or_404(id)
    if request.method == 'GET':
        units = Unit.query.all()
        departments = Department.query.all()
        product_names = ProductName.query.all()
        brands = Brand.query.all()
        origins = Origin.query.all()
        warehouses = Warehouse.query.all()
        return render_template('purchase/form.html', order=order, suppliers=units, departments=departments,
                               product_names=product_names, brands=brands, origins=origins, warehouses=warehouses, edit_mode=True)

    for item in order.items:
        inv = Inventory.query.filter_by(card_no=item.card_no).first()
        if inv and inv.status != 'in_stock':
            flash('该入库单的库存' + inv.card_no + '已被加工或出库消耗，请先删除下游单据', 'error')
            return redirect(url_for('purchase.detail', id=order.id))
        si = SalesItem.query.filter_by(card_no=item.card_no).first()
        if si:
            flash('该入库单的库存' + item.card_no + '已被出库使用，请先删除出库单', 'error')
            return redirect(url_for('purchase.detail', id=order.id))
        pd = ProcessDetail.query.filter_by(raw_card_no=item.card_no).first()
        if pd:
            flash('该入库单的库存' + item.card_no + '已被加工使用，请先删除加工单', 'error')
            return redirect(url_for('purchase.detail', id=order.id))

    for item in order.items:
        inv = Inventory.query.filter_by(card_no=item.card_no).first()
        if inv:
            db.session.delete(inv)
        TransactionLog.query.filter_by(card_no=item.card_no, order_no=order.order_no).delete()
    for item in order.items:
        db.session.delete(item)

    order.supplier_id = request.form.get('supplier_id', 0, type=int)
    order.dept_id = request.form.get('dept_id', 0, type=int)
    order.remark = request.form.get('remark', '')
    order_date_str = request.form.get('order_date', '')
    if order_date_str:
        order.order_date = date_type.fromisoformat(order_date_str)

    card_nos = request.form.getlist('card_no[]')
    product_names = request.form.getlist('product_name[]')
    brands_list = request.form.getlist('brand[]')
    origins_list = request.form.getlist('origin[]')
    specs = request.form.getlist('spec[]')
    qtys = request.form.getlist('qty[]')
    weights = request.form.getlist('weight[]')
    unit_prices = request.form.getlist('unit_price[]')
    remarks = request.form.getlist('remark[]')
    warehouses = request.form.getlist('warehouse[]')

    total_qty = 0
    total_weight = 0
    total_amount = 0
    all_warehouses = {w.id: w.name for w in Warehouse.query.all()}
    _wh_id = request.form.get('warehouse_id', type=int)
    wh_name = all_warehouses.get(_wh_id, '') if _wh_id else ''

    for i in range(len(card_nos)):
        unit_price = float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else 0
        amount = float(weights[i]) * unit_price
        pi = PurchaseItem(
            order_id=order.id,
            card_no=card_nos[i],
            product_name=product_names[i],
            brand=brands_list[i],
            origin=origins_list[i],
            spec=specs[i],
            qty=int(qtys[i]),
            weight=float(weights[i]),
            unit_price=unit_price,
            amount=amount,
            warehouse=wh_name,
            remark=remarks[i] if i < len(remarks) else ''
        )
        db.session.add(pi)

        inv = Inventory(
            card_no=card_nos[i],
            product_name=product_names[i],
            brand=brands_list[i],
            origin=origins_list[i],
            spec=specs[i],
            warehouse_id=_wh_id,
            qty=int(qtys[i]),
            weight=float(weights[i]),
            dept_id=order.dept_id,
            status='in_stock'
        )
        db.session.add(inv)

        tlog = TransactionLog(
            card_no=card_nos[i],
            type='in',
            product_name=product_names[i],
            spec=specs[i],
            brand=brands_list[i],
            origin=origins_list[i],
            order_no=order.order_no,
            weight_before=0,
            weight_after=float(weights[i])
        )
        db.session.add(tlog)

        total_qty += int(qtys[i])
        total_weight += float(weights[i])
        total_amount += amount

    order.total_qty = total_qty
    order.total_weight = total_weight
    order.total_amount = total_amount
    db.session.commit()
    flash('入库单更新成功', 'success')
    return redirect(url_for('purchase.detail', id=order.id))


@purchase_bp.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete(id):
    order = PurchaseOrder.query.get_or_404(id)
    for item in order.items:
        inv = Inventory.query.filter_by(card_no=item.card_no).first()
        if inv and inv.status != 'in_stock':
            flash('该入库单的库存' + inv.card_no + '已被加工或出库消耗，请先删除下游单据', 'error')
            return redirect(url_for('purchase.index'))
        si = SalesItem.query.filter_by(card_no=item.card_no).first()
        if si:
            flash('该入库单的库存' + item.card_no + '已被出库使用，请先删除出库单', 'error')
            return redirect(url_for('purchase.index'))
        pd = ProcessDetail.query.filter_by(raw_card_no=item.card_no).first()
        if pd:
            flash('该入库单的库存' + item.card_no + '已被加工使用，请先删除加工单', 'error')
            return redirect(url_for('purchase.index'))
    for item in order.items:
        inv = Inventory.query.filter_by(card_no=item.card_no).first()
        if inv:
            db.session.delete(inv)
        TransactionLog.query.filter_by(card_no=item.card_no, order_no=order.order_no).delete()
    db.session.delete(order)
    db.session.commit()
    flash('入库单已删除', 'success')
    return redirect(url_for('purchase.index'))


@purchase_bp.route('/template')
@login_required
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '入库模板'
    headers = ['供应商', '销售部门', '品名', '卡号', '牌号', '产地', '仓库', '规格', '件数', '吨位', '备注']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='入库模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@purchase_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'GET':
        return redirect(url_for('purchase.index'))

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('请上传Excel文件（.xlsx或.xls）', 'error')
        return redirect(url_for('purchase.index'))

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            flash('Excel文件中没有数据行', 'error')
            return redirect(url_for('purchase.index'))

        supplier_names = {s.name for s in Unit.query.all()}
        dept_names = {d.name for d in Department.query.all()}
        pn_names = {p.name for p in ProductName.query.all()}
        brand_names = {b.name for b in Brand.query.all()}
        origin_names = {o.name for o in Origin.query.all()}

        preview_rows = []
        for row in rows[:500]:
            if not row or not any(row):
                continue
            supplier = str(row[0] or '').strip()
            dept = str(row[1] or '').strip()
            product_name_val = str(row[2] or '').strip()
            card_no = str(row[3] or '').strip()
            brand_val = str(row[4] or '').strip()
            origin_val = str(row[5] or '').strip()
            warehouse_val = str(row[6] or '').strip()
            spec_val = str(row[7] or '').strip()
            qty = int(row[8] or 0)
            weight = float(row[9] or 0)
            remark = str(row[10] or '').strip()

            errors = {}
            if supplier not in supplier_names:
                errors['supplier'] = True
            if dept not in dept_names:
                errors['dept'] = True
            if product_name_val not in pn_names:
                errors['product_name'] = True
            if brand_val not in brand_names:
                errors['brand'] = True
            if origin_val not in origin_names:
                errors['origin'] = True
            if not card_no:
                errors['card_no'] = True
            if weight <= 0:
                errors['weight'] = True
            if qty < 0:
                errors['qty'] = True

            preview_rows.append({
                'supplier': supplier,
                'dept': dept,
                'product_name': product_name_val,
                'card_no': card_no,
                'brand': brand_val,
                'origin': origin_val,
                'warehouse': warehouse_val,
                'spec': spec_val,
                'qty': qty,
                'weight': weight,
                'remark': remark,
                'errors': errors,
            })

        session['import_preview_data'] = preview_rows
        return redirect(url_for('purchase.import_preview'))

    except Exception as e:
        flash(f'导入失败：{str(e)}', 'error')
        return redirect(url_for('purchase.index'))


@purchase_bp.route('/import/preview')
@login_required
def import_preview():
    preview_rows = session.get('import_preview_data')
    if not preview_rows:
        flash('没有待预览的导入数据', 'error')
        return redirect(url_for('purchase.index'))

    supplier_names = sorted(s.name for s in Unit.query.all())
    dept_names = sorted(d.name for d in Department.query.all())
    pn_names = sorted(p.name for p in ProductName.query.all())
    brand_names = sorted(b.name for b in Brand.query.all())
    origin_names = sorted(o.name for o in Origin.query.all())

    return render_template('purchase/import_preview.html',
                           rows=preview_rows,
                           total=len(preview_rows),
                           supplier_names=supplier_names,
                           dept_names=dept_names,
                           pn_names=pn_names,
                           brand_names=brand_names,
                           origin_names=origin_names)


@purchase_bp.route('/import/delete-row', methods=['POST'])
@login_required
def import_delete_row():
    preview_rows = session.get('import_preview_data')
    row_idx = request.form.get('row_idx', type=int)
    if preview_rows and row_idx is not None and 0 <= row_idx < len(preview_rows):
        preview_rows.pop(row_idx)
        session['import_preview_data'] = preview_rows
    return redirect(url_for('purchase.import_preview'))


@purchase_bp.route('/import/confirm', methods=['POST'])
@login_required
def import_confirm():
    preview_rows = session.get('import_preview_data')
    if not preview_rows:
        flash('没有待确认的导入数据', 'error')
        return redirect(url_for('purchase.index'))

    supplier_names = {s.name: s.id for s in Unit.query.all()}
    dept_names = {d.name: d.id for d in Department.query.all()}
    pn_names = {p.name for p in ProductName.query.all()}
    brand_names = {b.name for b in Brand.query.all()}
    origin_names = {o.name for o in Origin.query.all()}

    updated_rows = []
    errors_found = False
    error_rows = []
    for i, row in enumerate(preview_rows):

        row['supplier'] = request.form.get(f'supplier_{i}', row['supplier']).strip()
        row['dept'] = request.form.get(f'dept_{i}', row['dept']).strip()
        row['product_name'] = request.form.get(f'product_name_{i}', row['product_name']).strip()
        row['card_no'] = request.form.get(f'card_no_{i}', row['card_no']).strip()
        row['brand'] = request.form.get(f'brand_{i}', row['brand']).strip()
        row['origin'] = request.form.get(f'origin_{i}', row['origin']).strip()
        row['spec'] = request.form.get(f'spec_{i}', row['spec']).strip()
        row['warehouse'] = request.form.get(f'warehouse_{i}', row.get('warehouse', '')).strip()
        try:
            row['qty'] = int(request.form.get(f'qty_{i}', row['qty']))
        except (ValueError, TypeError):
            row['qty'] = 0
        try:
            row['weight'] = float(request.form.get(f'weight_{i}', row['weight']))
        except (ValueError, TypeError):
            row['weight'] = 0.0
        row['remark'] = request.form.get(f'remark_{i}', row['remark']).strip()

        errors = {}
        if row['supplier'] not in supplier_names:
            errors['supplier'] = True
        if row['dept'] not in dept_names:
            errors['dept'] = True
        if row['product_name'] not in pn_names:
            errors['product_name'] = True
        if row['brand'] not in brand_names:
            errors['brand'] = True
        if row['origin'] not in origin_names:
            errors['origin'] = True
        if not row['card_no']:
            errors['card_no'] = True
        if row['weight'] <= 0:
            errors['weight'] = True
        if row['qty'] < 0:
            errors['qty'] = True
        row['errors'] = errors

        if errors:
            errors_found = True
            error_rows.append(str(i + 1))
        updated_rows.append(row)

    if errors_found:
        session['import_preview_data'] = updated_rows
        flash(f'第 {",".join(error_rows)} 行字段未匹配基础数据，请修正后重新提交', 'error')
        return redirect(url_for('purchase.import_preview'))

    sorted_rows = sorted(updated_rows, key=lambda r: (r['supplier'], r['dept']))
    imported = 0
    wh_map = {w.name: w.id for w in Warehouse.query.all()}
    for (s_name, d_name), group_iter in groupby(sorted_rows, key=lambda r: (r['supplier'], r['dept'])):
        items_data = []
        for r in group_iter:
            wh_id = wh_map.get(r.get('warehouse', ''))
            items_data.append({
                'card_no': r['card_no'],
                'product_name': r['product_name'],
                'brand': r['brand'],
                'origin': r['origin'],
                'spec': r['spec'],
                'qty': r['qty'],
                'weight': r['weight'],
                'warehouse': r.get('warehouse', ''),
                'warehouse_id': wh_id if wh_id else None,
                'remark': r['remark'],
            })
        order_data = {
            'supplier_id': supplier_names[s_name],
            'dept_id': dept_names[d_name],
            'remark': 'Excel导入',
            'operator_id': current_user.id,
        }
        create_purchase_order(order_data, items_data)
        imported += 1

    session.pop('import_preview_data', None)
    flash(f'成功导入 {imported} 个入库单（共{len(updated_rows)}条明细）', 'success')
    return redirect(url_for('purchase.index'))


@purchase_bp.route('/detail-list')
@login_required
def detail_list():
    page = request.args.get('page', 1, type=int)
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    supplier_id = request.args.get('supplier_id', 0, type=int)
    dept_id = request.args.get('dept_id', 0, type=int)
    card_no = request.args.get('card_no', '').strip()
    spec = request.args.get('spec', '').strip()
    remark = request.args.get('remark', '').strip()
    order_remark = request.args.get('order_remark', '').strip()
    search_submitted = request.args.get('search_submitted', '')

    query = PurchaseItem.query.join(PurchaseOrder)
    if date_start:
        query = query.filter(PurchaseOrder.order_date >= date_start)
    if date_end:
        query = query.filter(PurchaseOrder.order_date <= date_end)
    if supplier_id:
        query = query.filter(PurchaseOrder.supplier_id == supplier_id)
    if dept_id:
        query = query.filter(PurchaseOrder.dept_id == dept_id)
    if card_no:
        query = query.filter(PurchaseItem.card_no.contains(card_no))
    if spec:
        query = query.filter(PurchaseItem.spec.contains(spec))
    if remark:
        query = query.filter(PurchaseItem.remark.contains(remark))
    if order_remark:
        query = query.filter(PurchaseOrder.remark.contains(order_remark))

    is_export = request.args.get('export', '') == '1'
    default_date_start = (date_type.today() - timedelta(days=30)).isoformat()
    default_date_end = date_type.today().isoformat()
    if not search_submitted and not is_export:
        units = Unit.query.all()
        departments = Department.query.all()
        return render_template('purchase/detail_list.html', pagination=None, suppliers=units, departments=departments,
                               search_submitted=False, default_date_start=default_date_start, default_date_end=default_date_end)

    pagination = query.order_by(PurchaseOrder.order_date.desc(), PurchaseItem.id.desc()).paginate(
        page=page, per_page=current_app.config['PER_PAGE'], error_out=False
    )

    if is_export:
        all_items = query.order_by(PurchaseOrder.order_date.desc(), PurchaseItem.id.desc()).all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '入库明细'
        headers = ['单号', '业务日期', '供应商', '销售部门', '卡号', '品名', '牌号', '产地', '规格', '件数', '吨位', '备注']
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        for item in all_items:
            ws.append([
                item.order.order_no,
                item.order.order_date.strftime('%Y-%m-%d') if item.order.order_date else '',
                item.order.supplier.name if item.order.supplier else '',
                item.order.dept.name if item.order.dept else '',
                item.card_no,
                item.product_name,
                item.brand,
                item.origin,
                item.spec,
                item.qty,
                float(item.weight),
                item.remark or '',
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='入库明细.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    units = Unit.query.all()
    departments = Department.query.all()
    total_qty_sum = sum(i.qty for i in pagination.items)
    total_weight_sum = sum(float(i.weight) for i in pagination.items)
    total_amount_sum = sum(float(i.amount) for i in pagination.items)
    return render_template('purchase/detail_list.html', pagination=pagination, suppliers=units, departments=departments,
                           search_submitted=True, total_qty_sum=total_qty_sum, total_weight_sum=total_weight_sum,
                           total_amount_sum=total_amount_sum, default_date_start=default_date_start, default_date_end=default_date_end)
